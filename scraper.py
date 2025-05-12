import asyncio
import pandas as pd
import os
from openpyxl import load_workbook
from playwright.async_api import async_playwright

async def scroll_to_load_all(page):
    previous_height = 0
    while True:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(1.5)
        current_height = await page.evaluate("document.body.scrollHeight")
        if current_height == previous_height:
            break
        previous_height = current_height

async def scrape_streaming():
    filename = "Faculty_Academic_Interests.xlsx"
    if not os.path.exists(filename):
        pd.DataFrame(columns=["Name", "Profile Link", "All Text"]).to_excel(filename, index=False)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        yield "🔎 Navigating to faculty list page..."
        await page.goto("https://www.torontomu.ca/engineering-architectural-science/about/faculty-search/")
        await page.wait_for_selector("a.name", timeout=10000)

        yield "📜 Scrolling to load all profiles..."
        await scroll_to_load_all(page)

        faculty_links = page.locator("a.name")
        names = await faculty_links.all_inner_texts()
        hrefs = await faculty_links.evaluate_all("els => els.map(el => el.href)")

        yield f"👨‍🏫 Found {len(names)} faculty members!"

        for name, link in zip(names, hrefs):
            try:
                yield f"🔍 Scraping {name}..."
                prof_page = await browser.new_page()
                await prof_page.goto(link)
                await prof_page.wait_for_load_state("domcontentloaded")

                all_text = await prof_page.locator("body").inner_text()
                new_entry = pd.DataFrame([[name, link, all_text]], columns=["Name", "Profile Link", "All Text"])

                book = load_workbook(filename)
                start_row = book.active.max_row

                with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                    new_entry.to_excel(writer, index=False, header=False, startrow=start_row)

                await prof_page.close()
                yield f"✅ Saved data for {name}."
            except Exception as e:
                yield f"❌ Error with {name}: {e}"

        await browser.close()
        yield "✅ Scraping complete!"
